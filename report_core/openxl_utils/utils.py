#!/usr/local/bin/python3
# -*- coding: utf-8 -*-

"""
@File    : utils.py
@Author  : Link
@Time    : 2022/7/31 20:15
@Mark    : 使用多进程来操作
"""
import os

import win32api
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

import pandas as pd

from common.app_variable import GlobalVariable
from common.func import tid_maker


class OpenXl:
    CenterAlign = Alignment(horizontal='center', vertical='center')
    RightAlign = Alignment(horizontal='right', vertical='center')
    LeftAlign = Alignment(horizontal='left', vertical='center', wrap_text=True)

    BlueFill = PatternFill('solid', fgColor='00B050')
    YellowFill = PatternFill('solid', fgColor='FFFF00')
    RedFill = PatternFill('solid', fgColor='FF0000')
    GreenFill = PatternFill('solid', fgColor='EBF1DE')

    CommentFont = Font(name="等线", size=9, color="FF0000")
    SpecialFont = Font(name="微软雅黑", size=9, color="0070C0")
    TitleFont = Font(name="微软雅黑", size=10, bold=True)
    ColumnFont = Font(name="等线", size=11, bold=True)
    TextFont = Font(name="等线", size=9)

    Thin = Side(border_style="thin", color="000000")
    TextBorder = Border(top=Thin, left=Thin, right=Thin, bottom=Thin)

    @staticmethod
    def excel_limit_run(summary_df: pd.DataFrame, limit_df: pd.DataFrame):
        """
        前四列固定, 从第二行开始写入
        """
        wb = Workbook()

        sheet_1 = wb.create_sheet("Limit")

        row_head = ["ID", "LOT_ID", "SBLOT_ID", "WAFER_ID", "TEST_COD", "FLOW_ID", "PART_TYP", "JOB_NAM"]
        column_head = ["TEST_ID", "TEXT", "UNITS"]
        df_limit = limit_df[column_head].drop_duplicates(subset=None, keep='first')

        for row, each in enumerate(row_head):
            t_row = row + 1
            sheet_1.cell(t_row, 2).value = each
            sheet_1.cell(t_row, 2).font = OpenXl.TitleFont
            sheet_1.cell(t_row, 2).fill = OpenXl.BlueFill

        title_row = len(row_head) + 1
        head_row = title_row + 1

        now_column = 1
        for each in column_head:
            for i in range(len(df_limit)):
                limit_data = df_limit.iloc[i]
                sheet_1.cell(head_row + i, now_column).value = limit_data[each]
                sheet_1.cell(head_row + i, now_column).font = OpenXl.SpecialFont
                sheet_1.cell(head_row + i, now_column).border = OpenXl.TextBorder
            sheet_1.cell(title_row, now_column).value = each
            sheet_1.cell(title_row, now_column).alignment = OpenXl.CenterAlign
            sheet_1.cell(title_row, now_column).font = OpenXl.SpecialFont
            sheet_1.cell(title_row, now_column).fill = OpenXl.YellowFill
            sheet_1.cell(title_row, now_column).border = OpenXl.TextBorder
            now_column += 1

        summary_info = summary_df.to_dict(orient='records')

        for index, lot in enumerate(summary_info):

            for row, each in enumerate(row_head):
                start_row = row + 1
                sheet_1.merge_cells(start_row=start_row, start_column=now_column, end_row=start_row,
                                    end_column=now_column + 1)
                sheet_1.cell(start_row, now_column).value = lot[each]
                sheet_1.cell(start_row, now_column).alignment = OpenXl.CenterAlign
                sheet_1.cell(start_row, now_column).font = OpenXl.TitleFont
                sheet_1.cell(start_row, now_column).fill = OpenXl.BlueFill
                # sheet_1.cell(start_row, now_column).border = OpenXl.TextBorder

            for each in ["LO_LIMIT", "HI_LIMIT"]:
                temp_lot = limit_df[limit_df.ID == lot["ID"]]
                temp_lot = temp_lot.set_index("TEXT")

                sheet_1.cell(title_row, now_column).value = each
                sheet_1.cell(title_row, now_column).alignment = OpenXl.CenterAlign
                sheet_1.cell(title_row, now_column).font = OpenXl.SpecialFont
                sheet_1.cell(title_row, now_column).fill = OpenXl.YellowFill
                sheet_1.cell(title_row, now_column).border = OpenXl.TextBorder

                for i in range(len(df_limit)):
                    limit_data = df_limit.iloc[i]
                    if limit_data.TEXT not in temp_lot.index:
                        continue
                    lot_limit_data = temp_lot.loc[limit_data.TEXT][each]  # 2022-1-20: CP25片用一个lot
                    sheet_1.cell(head_row + i, now_column).value = lot_limit_data
                    sheet_1.cell(head_row + i, now_column).alignment = OpenXl.CenterAlign
                    sheet_1.cell(head_row + i, now_column).font = OpenXl.TextFont
                    sheet_1.cell(head_row + i, now_column).border = OpenXl.TextBorder
                    if index != 0:
                        if sheet_1.cell(head_row + i, now_column).value != sheet_1.cell(head_row + i,
                                                                                        now_column - 2).value:
                            sheet_1.cell(head_row + i, now_column).fill = OpenXl.RedFill

                now_column += 1

        sheet_1.freeze_panes = "D{}".format(head_row)
        sheet_name = wb.worksheets[0]
        wb.remove(sheet_name)
        try:
            save_path = os.path.join(GlobalVariable.LIMIT_PATH, 'limit.xlsx')
            wb.save(save_path)
        except:
            save_path = os.path.join(GlobalVariable.LIMIT_PATH, 'limit_{}.xlsx'.format(tid_maker()))
            wb.save(save_path)
        win32api.ShellExecute(0, 'open', save_path, '', '', 1)

    @staticmethod
    def excel_lot_compare_run(lot_df: pd.DataFrame, cpk_df: pd.DataFrame):
        """
        简单的数据, 使用快速透视的方法来写入到Excel中
        先尝试删除本地的数据, 如果删除不了, 就新建一个数据
        :return:
        """
        save_path = os.path.join(GlobalVariable.LIMIT_PATH, "compare.xlsx")
        try:
            if os.path.exists(save_path):
                os.remove(save_path)
        except:
            save_path = os.path.join(GlobalVariable.LIMIT_PATH, "compare_{}.xlsx".format(tid_maker()))
        df = pd.merge(lot_df, cpk_df, on="ID")
        with pd.ExcelWriter(save_path) as writer:
            avg = df.pivot_table(index='TEXT',  # 透视的行，分组依据
                                 columns=['SUB_CON', 'PART_TYP', 'JOB_NAM', 'TEST_COD', 'LOT_ID', 'SBLOT_ID',
                                          'WAFER_ID', 'QTY'],
                                 values='AVG',  # 值
                                 aggfunc='sum'  # 聚合函数
                                 )
            std = df.pivot_table(index='TEXT',  # 透视的行，分组依据
                                 columns=['SUB_CON', 'PART_TYP', 'JOB_NAM', 'TEST_COD', 'LOT_ID', 'SBLOT_ID',
                                          'WAFER_ID', 'QTY'],
                                 values='STD',  # 值
                                 aggfunc='sum'  # 聚合函数
                                 )
            cpk = df.pivot_table(index='TEXT',  # 透视的行，分组依据
                                 columns=['SUB_CON', 'PART_TYP', 'JOB_NAM', 'TEST_COD', 'LOT_ID', 'SBLOT_ID',
                                          'WAFER_ID', 'QTY'],
                                 values='CPK',  # 值
                                 aggfunc='sum'  # 聚合函数
                                 )
            avg.to_excel(writer, 'AVG')
            std.to_excel(writer, 'STD')
            cpk.to_excel(writer, 'CPK')
        win32api.ShellExecute(0, 'open', save_path, '', '', 1)
